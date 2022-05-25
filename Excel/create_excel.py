from openpyxl import load_workbook
from openpyxl import Workbook
import json

if __name__ == "__main__":
    pass

def main():
    print('hello')
    #js = json.load("basic_condition.json")
    wb = load_workbook('create_sample.xlsx')
    #ws = wb[0]
    x = 0
    ws = wb['Sheet']

    ws.append(['エア圧', 'distance', 'velocity'])
    ws['A2'] = datetime.
    wb.save('create_sample.xlsx')

main()